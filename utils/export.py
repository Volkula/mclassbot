import csv
import io
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.orm import Session
from database.models import Event, Registration, User


def export_registrations_to_csv(db: Session, event_id: int) -> str:
    """Экспорт регистраций на событие в CSV"""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise ValueError("Событие не найдено")
    
    registrations = db.query(Registration).filter(Registration.event_id == event_id).all()
    
    output = io.StringIO()
    
    # Заголовки
    headers = ["ID", "Telegram ID", "Имя", "Ник", "Ссылка на Telegram", "Дата регистрации"]
    for field in sorted(event.fields, key=lambda x: x.order):
        headers.append(field.field_name)
    
    writer = csv.writer(output)
    writer.writerow(headers)
    
    # Данные
    for reg in registrations:
        user = db.query(User).filter(User.telegram_id == reg.user_telegram_id).first()
        user_name = user.full_name if user and user.full_name else ""
        user_username = user.username if user and user.username else ""
        
        # Формируем ссылку на Telegram
        if user_username:
            telegram_link = f"https://t.me/{user_username}"
        else:
            telegram_link = f"https://t.me/user{reg.user_telegram_id}"
        
        row = [
            reg.id,
            reg.user_telegram_id,
            user_name,
            user_username,
            telegram_link,
            reg.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]
        for field in sorted(event.fields, key=lambda x: x.order):
            value = reg.data_json.get(field.field_name, "")
            row.append(str(value))
        writer.writerow(row)
    
    return output.getvalue()


def export_registrations_to_excel(db: Session, event_id: int) -> bytes:
    """Экспорт регистраций на событие в Excel"""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise ValueError("Событие не найдено")
    
    registrations = db.query(Registration).filter(Registration.event_id == event_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрации"
    
    # Заголовки
    headers = ["ID", "Telegram ID", "Имя", "Ник", "Ссылка на Telegram", "Дата регистрации"]
    for field in sorted(event.fields, key=lambda x: x.order):
        headers.append(field.field_name)
    
    ws.append(headers)
    
    # Стиль для заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    row_num = 2
    for reg in registrations:
        user = db.query(User).filter(User.telegram_id == reg.user_telegram_id).first()
        user_name = user.full_name if user and user.full_name else ""
        user_username = user.username if user and user.username else ""
        
        # Формируем ссылку на Telegram
        if user_username:
            telegram_link = f"https://t.me/{user_username}"
        else:
            telegram_link = f"https://t.me/user{reg.user_telegram_id}"
        
        row = [
            reg.id,
            reg.user_telegram_id,
            user_name,
            user_username,
            telegram_link,
            reg.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]
        for field in sorted(event.fields, key=lambda x: x.order):
            value = reg.data_json.get(field.field_name, "")
            row.append(str(value))
        
        ws.append(row)
        
        # Добавляем гиперссылку для столбца со ссылкой (столбец E, индекс 5)
        link_cell = ws.cell(row=row_num, column=5)
        link_cell.hyperlink = telegram_link
        link_cell.font = Font(color="0000FF", underline="single")
        link_cell.value = telegram_link
        
        row_num += 1
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

